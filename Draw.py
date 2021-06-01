import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import matplotlib.colors as mcolors
from openpyxl import load_workbook
from openpyxl import Workbook
book=Workbook()
data=book.active
dfrow = load_workbook(filename="Result.xlsx",  data_only=True)
df = dfrow.active
def col(c,df):
    col=[]
    for i in range(2,51,4):
        col.append(df.cell(row=i,column=c).value)
    return col
def col10(c,df):
    col=[]
    for i in range(2,51,4):
        col.append(df.cell(row=i,column=c).value*10)
    return col
def col5(c,df):
    col=[]
    for i in range(2,51,4):
        col.append(df.cell(row=i,column=c).value*5)
    return col
n1 = col(1,df)
n = col10(1,df)
p = col5(1,df)
RA_time = col(2,df)
SA_time = col(3,df)
pksize = col(4,df)
Enc_key_time=col(5,df)
Enc_key_size=col(6,df)
Dec_key_time=col(7,df)
Dec_key_size=col(8,df)
Enc_time=col(9,df)
ciphertext_size=col(10,df)
Sanitize_time=col(11,df)
Sanitize_cipher_size= col(12,df)
decryptiontime=col(13,df)

fig, ((ax0),(ax3)) = plt.subplots(nrows=2, ncols=1,
                                    figsize=(8, 8))
#ax0.set_title('Registration time')
ax0.errorbar(n,RA_time,color='maroon', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6, label='RA_Gen time')
ax0.grid()
ax0.legend(loc=2,prop={'size': 12})
ax0.set_xlabel('Totall Number of Attributes')
ax0.set_ylabel('Time (ms)')


ax3.errorbar(p,ciphertext_size,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='purple', markersize=6,label='Ciphertext Size')
ax3.grid()
ax3.legend(loc=2,prop={'size': 12})
ax3.set_xlabel('Number of Attributes')
ax3.set_ylabel('Size (byte)')
#ax3.set_ylim([5.4,5.8])
plt.draw()
plt.savefig("performance.pdf", bbox_inches='tight')

'''

#ax1.set_title('Collateral size')
ax1.plot(n1,pksize,color='maroon', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6,label='PK size')
ax1.grid()
ax1.legend(loc=2,prop={'size': 12})
ax1.set_xlabel('Number of Attributes')
ax1.set_ylabel('Size (byte)')
#ax3.set_title('ciphertext Size')

#ax2.set_title('Speending time')
ax2.plot(n1,Enc_key_time,color='darkgreen', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='maroon', markersize=6,label='Enc_key time')
ax2.plot(n1,Dec_key_time,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6,label='Dec_key time')
ax2.grid()
ax2.legend(loc=2,prop={'size': 12})
ax2.set_xlabel('Number of Attributes')
ax2.set_ylabel('Time (sec)')

#ax4.set_title('Key Gen time')
ax4.errorbar(n1,Enc_time,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='r', markersize=6,label='Encryption time')
ax4.grid()
ax4.legend(loc=2,prop={'size': 12})
ax4.set_xlabel('Number of Attributes')
ax4.set_ylabel('Time (ms)')
#ax4.set_ylim([0,7])

#ax5.set_title('Verification time')
ax5.errorbar(n1,decryptiontime,color='b', linestyle='dashed', linewidth = 2,
         marker='o', markerfacecolor='g', markersize=6,label='Decryption time')
ax5.grid()
ax5.legend(loc=2,prop={'size': 12})
ax5.set_xlabel('Number of Attributes')
ax5.set_ylabel('Time (Sec)')
#ax5.set_ylim([0.1,0.5])
'''
